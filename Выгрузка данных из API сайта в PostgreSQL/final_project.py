# Импортируем необходимые библиотеки
import requests  # Для запроса данных их API
import logging  # Для логирования
from logging.handlers import TimedRotatingFileHandler
from pathlib import Path
import psycopg2  # Для работы с PostgreSQL
from webdav3.client import Client # Для работы с WebDAV
from openpyxl import load_workbook, Workbook # Для работы с xlsx-файлами
from datetime import datetime, timedelta   # Для работы с датой и временем
import ast  # для чтения "кода", чтобы быстро строку превратить в словарь
import os
import smtplib  # Для отправки по SMTP
from email.message import EmailMessage  # Для создания e-mail

# Считаем, что код запускается где-то ночью и записывает данные в БД за вчерашний день timedelta(days=1)
# Так как время на сервере в нулевом часовом поясе, нужно это учитывать. Поэтому можно сделать +3 часа к start и end
# Можно задать любое значение hours
hours = 3
yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
start_datetime = datetime.strptime(f'{yesterday} 00:00:00.000000', "%Y-%m-%d %H:%M:%S.%f") + timedelta(hours=hours)
end_datetime = datetime.strptime(f'{yesterday} 23:59:59.999999', "%Y-%m-%d %H:%M:%S.%f") + timedelta(hours=hours)
start = (start_datetime).strftime("%Y-%m-%d %H:%M:%S.%f")
end = (end_datetime).strftime("%Y-%m-%d %H:%M:%S.%f")

# Настройка логера:
logging.basicConfig(
    format='%(asctime)s %(levelname)s: %(message)s',
    level=logging.INFO,
    encoding='utf-8',
    filename=f"{Path(__file__).parent}/{yesterday}.log",
    filemode="a"
)
logging.info("Запуск программы...")
# Переменные
api_url = "https://b2b.itresume.ru/api/statistics"
params = {
    "client": "Skillfactory",
    "client_key": "M2MGWS",
    "start": start,
    "end": end
}

psql_address = "localhost"
psql_port = 5432
psql_db = "postgres"
psql_user = "postgres"
psql_pass = 123 # Можно вынести пароль в отдельную переменную, которую брать из файла (как для Я.Диска и E-mail сделано)


class ConnectDB:
    __instance = None  # Явное объявление приватного атрибута класса

    @staticmethod
    def get_instance():
        if not ConnectDB.__instance:
            ConnectDB()
        return ConnectDB.__instance

    def __init__(self):
        """Устанавливает соединение с БД"""
        if ConnectDB.__instance:
            logging.warning(
                f"Попытка создать второй экземпляр класса {self.__class__.__name__}")
            raise Exception("Этот класс является Singleton")
        else:
            self.connection = psycopg2.connect(
                host=psql_address,
                port=psql_port,
                database=psql_db,
                user=psql_user,
                password=psql_pass
            )
            ConnectDB.__instance = self
            logging.info(
                f'Соединение с БД "{psql_db}:{psql_port}" установлено')

    def commit(self):
        self.connection.commit()

    def rollback(self):
        self.connection.rollback()

    def close_instance(self):
        """Закрывает соединение с БД"""
        if ConnectDB.__instance:
            self.connection.close()
            logging.info(f'Соединение с БД "{psql_db}:{psql_port}" закрыто')


class APIClient:
    def __init__(self, url, params):
        self.url = url
        self.__params = params

    def fetch_data(self):
        """Загружает данные из API и возвращает их в структурированном виде."""
        try:
            response = requests.get(self.url, params=self.__params)
            self.status = response.status_code
            self.__json = response.json()
            logging.info('Данные из API получены')
        except requests.exceptions.HTTPError:
            logging.warning(f"HTTP Error: {self.status}")
            return self.status
        except requests.exceptions.RequestException:
            logging.warning(f"Request Error: {self.status}")
            return self.status
        res = []
        try:
            logging.info('Начата обработка данных')
            for el in self.__json:
                dict_res = {}
                dict_res['user_id'] = el['lti_user_id']
                if 'passback_params' in el:
                    if el['passback_params'] is not None:
                        d = ast.literal_eval((el['passback_params']))
                    else:
                        d = {}
                    for key in ('oauth_consumer_key', 'lis_result_sourcedid', 'lis_outcome_service_url'):
                        dict_res[key] = d[key] if key in d else ""
                else:
                    dict_res['oauth_consumer_key'] = ""
                    dict_res['lis_result_sourcedid'] = ""
                    dict_res['lis_outcome_service_url'] = ""
                dict_res['is_correct'] = el['is_correct']
                dict_res['attempt_type'] = el['attempt_type']
                dict_res['created_at'] = el['created_at']
                for k, v in dict_res.items():
                    if v == "":
                        logging.warning(
                            f'Обнаружено пустое значение в ключе {k}')
                res.append(dict_res)
            logging.info('Обработка данных закончена')
            return res
        except KeyError:
            logging.error('Ошибка ключа при формировании словаря')
            return []

class YandexDisk:
    def __init__(self):
        #Сохраню пароль в отдельном файле, который не выгружается на гит
        with open('password_yandex', 'r') as f:
            self.__password_xlsx = f.read()

        self.options = {
            "webdav_hostname": "https://webdav.yandex.ru",
            "webdav_login": "kolcharma@yandex.ru",
            "webdav_password": self.__password_xlsx,
        }
        
    def report(self, filename):
        self.filename = filename
        client = Client(self.options)
        if not client.check("python"):
            client.mkdir("python")
        client.upload(f'python/Отчёт.xlsx', filename)

# Получаем данные из API
data = APIClient(api_url, params).fetch_data()

db_connection = ConnectDB.get_instance()  # Соединяемся с БД
# Таблица создана заранее. Файл "create_table_users.sql"
try:
    cursor = db_connection.connection.cursor()  # Устанавливаем курсор
    query = """
    INSERT INTO users (
        user_id, 
        oauth_consumer_key, 
        lis_result_sourcedid, 
        lis_outcome_service_url, 
        is_correct, 
        attempt_type, 
        created_at
        ) 
        VALUES (
        %(user_id)s, 
        %(oauth_consumer_key)s, 
        %(lis_result_sourcedid)s,
        %(lis_outcome_service_url)s,
        %(is_correct)s,
        %(attempt_type)s,
        %(created_at)s
        )
    """
    cursor.executemany(query, data)  # Передаём список словарей для вставки многих строк
    db_connection.commit()  # "Подтверждаем" вставку
    logging.info(f'Данные успешно занесены в БД')
except Exception as e:
    logging.error(f"Неожиданная ошибка: {e}")
    if db_connection:
        # "Отменяем" вставку
        db_connection.rollback()
finally:
    if db_connection:
        # Отсоединяемся от БД
        db_connection.close_instance()

# Удаление лог-файлов старше 3х дней
def delete_old_log(days=3, folder='.'):
    cutoff = datetime.now() - timedelta(days=days)
    for filename in os.listdir(path):
        file_path = os.path.join(folder, filename)
        file_date = datetime.fromtimestamp(os.path.getmtime(file_path))
        condition = filename.lower().endswith('.log') and file_date <= cutoff
        if condition:
            os.remove(file_path)
            logging.info(f'Удалён лог-файл {filename}')

file = os.path.abspath(__file__) # Получаем путь к файлу .py
path = os.path.dirname(file)    # Берём только директорию
delete_old_log(days=3, folder=path) # Удаляем логи старше 3х дней от текущего момента

logging.info("Формируется отчёт в формате xlsx")

def cnt_data(data):
    cnt_users = set()
    cnt_attempts = 0
    cnt_success_attempts = 0
    for row in data:
        cnt_users.add(row['user_id'])
        # Количество попыток считаем, без учёта верных решений
        if (row['is_correct'] == 0 and row['attempt_type'] == 'submit') or row['attempt_type'] == 'run':
            cnt_attempts += 1
        if row['is_correct'] == 1 and row['attempt_type'] == 'submit':
            cnt_success_attempts += 1

    return {'date': yesterday, 'count_unique_users': len(cnt_users), 'count_attempts': cnt_attempts, 'count_success_attempts': cnt_success_attempts}

report_data = cnt_data(data)

file = os.path.abspath(__file__) # Получаем путь к файлу .py
path = os.path.dirname(file) 
report_name = f"{path}/Отчёт.xlsx"
if not os.path.exists(report_name):
    logging.warning(f'Отсутствует файл отчёта {report_name}. Создаём его...')
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    # Записываем заголовки
    headers = [
        "Дата",
        "Количество уникальных пользователей",
        "Количество попыток",
        "Количество успешных попыток"
    ]
    ws.append(headers)
    # Сохраняем файл
    wb.save(report_name)
    logging.info('Файл отчёта создан')

# Добавляем данные в отчёт
wb = load_workbook(report_name)
ws = wb.active
try:
    ws.append(list(report_data.values()))
    wb.save(report_name)
    logging.info('Данные успешно добавлены')
except Exception as e:
    logging.error(f'Ошибка добавления данных: {e}')

# Загружаем файл на Я.Диск
try:
    report = YandexDisk().report(report_name)
    logging.info("Отчёт в формате xlsx загружен на Я.Диск")
except Exception as e:
    logging.error(f"Ошибка загрузки на Я.Диск: {e}")

def send_yandex_email(sender_email, sender_password, receiver_email: str, subject: str, body):
    try:
        # Настройки SMTP Яндекс
        smtp_server = "smtp.yandex.ru"
        smtp_port = 465  # SSL
        
        # Создаем сообщение
        msg = EmailMessage()
        msg['From'] = sender_email
        msg['To'] = receiver_email
        msg['Subject'] = subject
        msg.set_content(body)
        
        # Отправляем письмо
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(sender_email, sender_password)
            server.send_message(msg)
        
        logging.info("E-mail успешно отправлен!")
    
    except Exception as e:
        logging.error(f"Ошибка отправки E-mail: {e}")

# Настройка параметров для электронной почты
yandex_email = "kolcharma@yandex.ru"
with open('password_mail', 'r') as f:
    app_password = f.read()
recipient = "kolcharma@gmail.com"

send_yandex_email(
    sender_email=yandex_email,
    sender_password=app_password,
    receiver_email=recipient,
    subject="Важное уведомление",
    body="""Здравствуйте!
Работа алгоритма закончена
С уважением,
Автоматизированная система"""
)

logging.info("Работа программы успешно завершена")